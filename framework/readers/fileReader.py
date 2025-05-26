import openpyxl
import pandas as pd
import re
from pathlib import Path
from framework.mobile.prints import text_print
from openpyxl.utils import column_index_from_string

class FileReader:
    _base_files_path = Path(__file__).parent.parent.parent / 'files' # Assumes 'files' is at project root

    @staticmethod
    def get_cell_value_from_excel(file_name, sheet_name, cell_name):
        excel_file_path = FileReader._base_files_path / file_name
        # print(f"[DEBUG] Excel file path being accessed for reading: {excel_file_path}") # Optional debug print
        try:
            if not excel_file_path.exists():
                text_print(f"Error: Excel file not found at {excel_file_path}")
                return None
            
            wb = openpyxl.load_workbook(excel_file_path)
            if sheet_name not in wb.sheetnames:
                text_print(f"Error: Sheet '{sheet_name}' not found in {excel_file_path}. Available sheets: {wb.sheetnames}")
                wb.close()
                return None
            
            sheet = wb[sheet_name]
            # openpyxl handles invalid cell names by raising a KeyError or similar, 
            # so direct check might be redundant unless specific pre-validation is needed.
            value = sheet[cell_name].value
            wb.close()
            return value
        except Exception as e:
            text_print(f"Error reading cell '{cell_name}' from sheet '{sheet_name}' in {excel_file_path}: {e}")
            if 'wb' in locals() and wb is not None:
                wb.close()
            return None

    @staticmethod
    def set_cell_value_in_excel(file_name, sheet_name, cell_name, value_to_enter):
        excel_file_path = FileReader._base_files_path / file_name
        # print(f"[DEBUG] Excel file path being accessed for writing: {excel_file_path}") # Optional debug print
        try:
            if not excel_file_path.exists():
                text_print(f"Error: Excel file not found at {excel_file_path}. Cannot write value.")
                return False

            wb = openpyxl.load_workbook(excel_file_path)

            if sheet_name not in wb.sheetnames:
                text_print(f"Info: Sheet '{sheet_name}' not found in {excel_file_path}. Creating new sheet.")
                sheet = wb.create_sheet(title=sheet_name)
            else:
                sheet = wb[sheet_name]
            
            sheet[cell_name] = value_to_enter
            wb.save(excel_file_path)  # Save the changes to the workbook
            wb.close()
            text_print(f"Successfully wrote '{value_to_enter}' to cell '{cell_name}' in sheet '{sheet_name}' of {excel_file_path}")
            return True
        except Exception as e:
            text_print(f"Error writing to cell '{cell_name}' in sheet '{sheet_name}' of {excel_file_path}: {e}")
            if 'wb' in locals() and wb is not None:
                wb.close()
            return False

        # df = pd.read_csv(csv_path)
        # text_print(f"CSV file found at {df}")
        # val = df.at[0, 'phone_number']
        # val = str(val)  # <- Convert it to string
    
    @staticmethod
    def get_cell_value_from_csv(file_name, row_index, column_name):
        csv_path = FileReader._base_files_path / file_name
        # text_print(f"CSV file found at {csv_path}")
        try:
            if not csv_path.exists():
                text_print(f"Error: CSV file not found at {csv_path}")
                return None
            df = pd.read_csv(csv_path)
            value = df.at[row_index, column_name]
            return str(value) if value is not None else None
        except Exception as e:
            text_print(f"Error reading CSV file: {e}")
            return None

    @staticmethod
    def read_csv_cell(filename, cellname='A1'):
        """
        Reads a value from a specific cell in a CSV file using Excel-style cell names.

        Args:
            filename (str): The name of the CSV file (relative to _base_files_path).
            cellname (str): The Excel-style cell name (e.g., 'A1', 'B2'). Defaults to 'A1'.

        Returns:
            str or None: The value of the cell as a string, or None if the cell is out of bounds.

        Raises:
            ValueError: If the cell name is invalid.
            FileNotFoundError: If the specified file does not exist.
            RuntimeError: For other file operation or data access errors.
        """
        file_path = FileReader._base_files_path / filename

        # Convert cell reference like "B2" → row 1, column 1
        match = re.match(r"([A-Za-z]+)([0-9]+)", cellname)
        if not match:
            raise ValueError(f"Invalid cell name: {cellname}")

        col_letters, row_number = match.groups()
        try:
            col_index = column_index_from_string(col_letters) - 1  # 0-based
            row_index = int(row_number) - 1  # 0-based
        except ValueError:
             raise ValueError(f"Invalid cell name format: {cellname}")

        try:
            # Read CSV
            df = pd.read_csv(file_path)

            # Safely return value if it exists
            if row_index < len(df.index) and col_index < len(df.columns):
                return str(df.iat[row_index, col_index])
            else:
                # Return None for out of bounds as per original logic
                return None

        except FileNotFoundError:
            raise FileNotFoundError(f"File not found: {file_path}")
        except IndexError as e:
             # Catch IndexError specifically if df.iat raises it for out of bounds
             # Although the if condition above should prevent this, adding for robustness
             raise RuntimeError(f"Error accessing cell {cellname} in file {file_path}: {e}")
        except Exception as e:
            # Catch any other potential errors during read or access
            raise RuntimeError(f"Error reading CSV file {file_path}: {e}")

    @staticmethod
    def write_csv_cell(filename, cellname, value):
        """
        Writes a value to a specific cell in a CSV file using Excel-style cell reference.

        Args:
            filename (str): The name of the CSV file.
            cellname (str): The Excel-style cell reference (e.g., 'A1', 'B2').
            value (str): The value to write to the cell.

        Raises:
            ValueError: If the cell name is invalid or the cell is out of bounds.
        """
        filepath = FileReader._base_files_path / filename

        # Convert cell reference like "B2" → row 1, column 1
        match = re.match(r"([A-Za-z]+)([0-9]+)", cellname)
        if not match:
            raise ValueError(f"Invalid cell name: {cellname}")

        col_letters, row_number = match.groups()
        col_index = column_index_from_string(col_letters) - 1  # 0-based
        row_index = int(row_number) - 1  # 0-based

        try:
            # Read CSV
            df = pd.read_csv(filepath)

            # Check if row and column are within bounds
            if row_index >= len(df.index) or col_index >= len(df.columns):
                 # Optionally, you could resize the DataFrame here if needed
                 # For now, raising an error for out of bounds write
                 raise ValueError(f"Cell {cellname} is out of bounds for CSV file {filename}")

            # Update the value
            df.iat[row_index, col_index] = value

            # Write the modified DataFrame back to CSV
            df.to_csv(filepath, index=False)

        except FileNotFoundError:
            raise FileNotFoundError(f"CSV file not found: {filepath}")
        except Exception as e:
            raise RuntimeError(f"Error writing to CSV file {filepath}: {e}")

    @staticmethod
    def set_cell_value_in_csv(file_name, row_index, column_name, value_to_enter):
        csv_path = FileReader._base_files_path / file_name
        try:
            if not csv_path.exists():
                text_print(f"Error: CSV file not found at {csv_path}")
                return False
            df = pd.read_csv(csv_path)
            df.at[row_index, column_name] = value_to_enter
            df.to_csv(csv_path, index=False)
            text_print(f"Updated {column_name} at row {row_index} to '{new_value}' in {file_name}")
            return True
        except Exception as e:
            text_print(f"Error writing to CSV file: {e}")
            return False# Example usage (optional, for testing directly in this file):


        